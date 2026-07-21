"""Build test fixtures from the source spreadsheet.

Extracts the Oncor rows from the 'All' sheet (raw) and the 'Filtered' sheet
(the curator's kept set) and writes them as canonical Plan dicts to
tests/fixtures/. The golden-file test uses these to prove our filter reproduces
the curator's honest judgment.

Run:  python scripts/build_fixtures.py
Requires: openpyxl  (dev dependency only)
"""
from __future__ import annotations

import glob
import json
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

import openpyxl  # noqa: E402

from htx.models import TDU_ONCOR, from_ptc_record  # noqa: E402

ROOT = os.path.join(os.path.dirname(__file__), "..")
FIXTURES = os.path.join(ROOT, "tests", "fixtures")


def _sheet_records(ws) -> list[dict]:
    header = [c.value for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        out.append(dict(zip(header, row)))
    return out


def main() -> None:
    matches = glob.glob(os.path.join(ROOT, "*.xlsx"))
    if not matches:
        sys.exit("No .xlsx source spreadsheet found in repo root.")
    src = matches[0]
    print(f"Reading {os.path.basename(src)}")
    wb = openpyxl.load_workbook(src, data_only=True)

    os.makedirs(FIXTURES, exist_ok=True)

    # Oncor-only fixtures drive the golden-file fidelity test.
    for sheet, out_name in (("All", "oncor_all.json"), ("Filtered", "oncor_filtered.json")):
        records = _sheet_records(wb[sheet])
        plans = [from_ptc_record(r) for r in records]
        oncor = [p.to_dict() for p in plans if p.tdu == TDU_ONCOR]
        path = os.path.join(FIXTURES, out_name)
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(oncor, fh, indent=0)
        print(f"  {sheet}: wrote {len(oncor)} Oncor plans -> {out_name}")

    # All-regions fixture drives the offline multi-region pipeline (and its tests).
    all_plans = [from_ptc_record(r) for r in _sheet_records(wb["All"])]
    all_dicts = [p.to_dict() for p in all_plans]
    with open(os.path.join(FIXTURES, "all_plans.json"), "w", encoding="utf-8") as fh:
        json.dump(all_dicts, fh, indent=0)
    print(f"  All: wrote {len(all_dicts)} plans (all regions) -> all_plans.json")


if __name__ == "__main__":
    main()
